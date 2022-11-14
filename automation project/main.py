import openpyxl

inventory_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inventory_file["Sheet1"]

products_per_supplier = {}
total_rev_supplier = {}
products_inv_less_10 = {}

for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_num = product_list.cell(product_row, 1).value
    inventory_price = product_list.cell(product_row, 5)

    # Calculation for number of products per supplier
    if supplier_name in products_per_supplier:
        current_number_products = products_per_supplier.get(supplier_name)
        products_per_supplier[supplier_name] = products_per_supplier[supplier_name] + 1
    else:
        print("Adding a new supplier ")
        products_per_supplier[supplier_name] = 1

    # Total inventory revenue per supplier
    if supplier_name in total_rev_supplier:
        current_total_value = total_rev_supplier.get(supplier_name)
        total_rev_supplier[supplier_name] = current_total_value + inventory * price
    else:
        total_rev_supplier[supplier_name] = inventory * price

    # Products with inventory less than 10 units
    if inventory < 10:
        products_inv_less_10[int(product_num)] = int(inventory)

    # Add a column with total inventory Price
    inventory_price.value = inventory * price

inventory_file.save("inventory_with_values.xlsx")

print(products_per_supplier)
print(total_rev_supplier)
print(products_inv_less_10)


